# -*- coding: utf-8 -*-
from __future__ import annotations
import pandas as pd
import numpy as np
import io
import unicodedata
import re
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

def _vn_normalize(s: str) -> str:
    """Chuẩn hóa chuỗi: bỏ dấu, viết thường, xóa khoảng trắng thừa."""
    if pd.isna(s) or s is None: return ""
    s = str(s).strip().lower()
    s = "".join(ch for ch in unicodedata.normalize("NFD", s) if unicodedata.category(ch) != "Mn")
    return re.sub(r"[\s\._\-]+", " ", s).strip()

def process_hd01(df: pd.DataFrame, store_name: str) -> pd.DataFrame:
    """
    Hàm xử lý file Excel HD01 thô. 
    Bảo tồn 100%: Chống lệch Index, Map chuẩn Doanh thu, Cột ép kiểu Text.
    """
    if df is None or df.empty: return pd.DataFrame()

    sub_idx = -1
    for i in range(min(20, len(df))):
        row_str = " ".join([str(x).lower() for x in df.iloc[i].values if pd.notna(x)])
        if 'seri' in row_str and 'số' in row_str and 'ngày' in row_str:
            sub_idx = i
            break

    if sub_idx == -1 or sub_idx == 0: 
        return pd.DataFrame()

    main_header = df.iloc[sub_idx - 1].copy()
    sub_header = df.iloc[sub_idx].copy()

    last_val = None
    for j in range(len(main_header)):
        val = main_header.iloc[j]
        if pd.notna(val) and str(val).strip() != "" and "Unnamed" not in str(val): 
            last_val = str(val).strip()
        elif last_val is not None: 
            main_header.iloc[j] = last_val

    combined_headers = []
    for j in range(len(main_header)):
        m_val = str(main_header.iloc[j]).strip() if pd.notna(main_header.iloc[j]) else ""
        s_val = str(sub_header.iloc[j]).strip() if pd.notna(sub_header.iloc[j]) else ""
        if m_val and s_val: combined_headers.append(f"{m_val} {s_val}")
        elif m_val: combined_headers.append(m_val)
        elif s_val: combined_headers.append(s_val)
        else: combined_headers.append(f"Cột_{j}")

    df_data = df.iloc[sub_idx + 1:].copy()
    df_data.columns = combined_headers

    df_data = df_data.dropna(subset=[df_data.columns[0]], how='all')
    df_data = df_data[~df_data.iloc[:, 0].astype(str).str.contains('STT|Tổng cộng', case=False, na=False)]

    df_data = df_data.reset_index(drop=True)

    if df_data.empty: return pd.DataFrame()

    def get_col_by_keywords(keywords: list) -> str | None:
        for col in df_data.columns:
            col_lower = str(col).lower()
            for kw in keywords:
                if kw.lower() in col_lower: return col
        return None

    col_mapping = {
        'Tên CHXD': None,
        'Ký hiệu': get_col_by_keywords(['ký hiệu hóa đơn', 'ky hieu hoa don', 'ký hiệu', 'ky hieu', 'seri']),
        'Số HĐ': get_col_by_keywords(['số hóa đơn', 'so hoa don', 'số hd', 'so hd', 'số hđ', 'số']),
        'Ngày hóa đơn': get_col_by_keywords(['ngày hóa đơn', 'ngay hoa don', 'ngày']),
        'Trạng thái HĐ': get_col_by_keywords(['trạng thái', 'trang thai']),
        'Loại HĐ': get_col_by_keywords(['loại hóa đơn', 'loại hoá đơn', 'loai hoa don', 'loại hđ', 'loai hd']),
        'Mã tra cứu': get_col_by_keywords(['mã tra cứu', 'ma tra cuu']),
        'Số GD': get_col_by_keywords(['số gd', 'so gd', 'số giao dịch', 'so giao dich']),
        'Mã khách hàng': get_col_by_keywords(['mã khách', 'ma khach']),
        'Tên khách hàng': get_col_by_keywords(['tên khách', 'ten khach']),
        'Mã số thuế': get_col_by_keywords(['mã số thuế', 'mst', 'ma so thue']),
        'Hàng hóa': get_col_by_keywords(['tên hàng', 'hàng hóa', 'hang hoa']),
        'ĐVT': get_col_by_keywords(['đvt', 'đơn vị tính', 'dvt']),
        'Số lượng': get_col_by_keywords(['số lượng', 'so luong']),
        'Đơn giá': get_col_by_keywords(['đơn giá', 'don gia']),
        'Thành tiền (chưa thuế)': get_col_by_keywords(['doanh thu (chưa thuế)', 'doanh thu', 'thành tiền', 'tien chua thue']),
        'Tiền thuế': get_col_by_keywords(['tiền thuế', 'thuế gtgt', 'tien thue']),
        'Tổng tiền thanh toán': get_col_by_keywords(['tổng tiền thanh toán', 'tổng tiền', 'tong tien'])
    }

    df_final = pd.DataFrame()
    df_final['Tên CHXD'] = [store_name] * len(df_data)

    for standardized_name, original_name in col_mapping.items():
        if standardized_name == 'Tên CHXD': continue
        if original_name and original_name in df_data.columns:
            if standardized_name in ['Số HĐ', 'Mã số thuế', 'Mã tra cứu', 'Số GD']:
                # ĐÃ SỬA: Loại bỏ dấu nháy đơn (') ở đầu vì BigQuery đã quản lý kiểu dữ liệu STRING rất tốt
                df_final[standardized_name] = df_data[original_name].apply(lambda x: str(x).strip() if pd.notna(x) and str(x).strip() else "")
            else:
                df_final[standardized_name] = df_data[original_name]
        else:
            df_final[standardized_name] = ""

    return df_final

def aggregate_hd01_data(dict_dfs: dict) -> io.BytesIO:
    if not dict_dfs: return None
    
    df_all = pd.concat(dict_dfs.values(), ignore_index=True)
    
    # BẢO VỆ MÃNG: Khởi tạo cột ảo nếu nó không tồn tại trong data để tránh lỗi KeyError
    if 'Loại HĐ' not in df_all.columns: df_all['Loại HĐ'] = ''
    if 'Mã số thuế' not in df_all.columns: df_all['Mã số thuế'] = ''

    # ÉP KIỂU SỐ HỌC
    for col in ['Thành tiền (chưa thuế)', 'Tiền thuế', 'Tổng tiền thanh toán', 'Số lượng']:
        if col in df_all.columns:
            df_all[col] = pd.to_numeric(df_all[col].astype(str).str.replace(',', '').str.replace(' ', ''), errors='coerce').fillna(0)

    # =====================================================================
    # LOGIC CẤY CỘT ẢO (VIRTUAL COLUMNS) CHO CHUYỂN THẲNG VÀ NỘI BỘ
    # =====================================================================
    # 1. Chuyển thẳng: Kiểm tra xem cột 'Loại HĐ' có chứa 'chuyển thẳng' không
    is_chuyen_thang = df_all['Loại HĐ'].astype(str).str.lower().str.contains('chuyển thẳng', na=False)
    df_all['SL_ChuyenThang'] = np.where(is_chuyen_thang, df_all['Số lượng'], 0)

    # 2. Nội bộ: Kiểm tra xem cột 'MST' có chứa '0600759399' không
    is_noi_bo = df_all['Mã số thuế'].astype(str).str.contains('0600759399', na=False)
    df_all['SL_NoiBo'] = np.where(is_noi_bo, df_all['Số lượng'], 0)

    target_products = ['Xăng RON95 Mức 3', 'Xăng E5 RON92 Mức 2', 'Dầu Điêzen 0,001S Mức 5', 'Dầu Điêzen 0,05S Mức 2']
    products = target_products + ['Mặt hàng khác']
    statuses = ['Hoàn thành', 'Thay thế', 'Điều chỉnh tăng', 'Điều chỉnh giảm', 'Bị thay thế', 'Bị điều chỉnh']
    valid_statuses_lower = [s.lower() for s in statuses]
    
    # ĐƯA 2 CHỈ SỐ MỚI VÀO TỪ ĐIỂN CẤU TRÚC BẢNG (Dynamic Extension)
    metrics = {
        'Tổng số lượng hóa đơn': 'invoice_count',
        'Tổng số tiền chưa thuế': 'Thành tiền (chưa thuế)',
        'Tiền thuế': 'Tiền thuế',
        'Tổng thanh toán': 'Tổng tiền thanh toán',
        'Giao dịch chuyển thẳng': 'SL_ChuyenThang',
        'Giao dịch nội bộ': 'SL_NoiBo'
    }

    if 'Hàng hóa' not in df_all.columns: df_all['Hàng hóa'] = ''
    df_all['Nhóm_Hàng'] = df_all['Hàng hóa'].apply(
        lambda x: x if pd.notna(x) and str(x).strip() in target_products else 'Mặt hàng khác'
    )

    if 'Trạng thái HĐ' not in df_all.columns: df_all['Trạng thái HĐ'] = ''
    df_all['Trạng_Thái_Lower'] = df_all['Trạng thái HĐ'].astype(str).str.strip().str.lower()
    
    # LỌC CÁC DÒNG HỢP LỆ THEO QUY TẮC
    df_all['Is_Valid'] = df_all['Trạng_Thái_Lower'].isin(valid_statuses_lower)
    df_valid = df_all[df_all['Is_Valid']]

    # THUẬT TOÁN ĐẾM VÀ TÍNH TỔNG SIÊU TỐC (ĐÃ BỔ SUNG 2 CỘT ẢO)
    agg_prod = df_valid.groupby(['Tên CHXD', 'Nhóm_Hàng']).agg(
        So_Luong_Dong=('Tên CHXD', 'size'),
        Tien_Chua_Thue=('Thành tiền (chưa thuế)', 'sum'),
        Tien_Thue=('Tiền thuế', 'sum'),
        Tong_Thanh_Toan=('Tổng tiền thanh toán', 'sum'),
        SL_ChuyenThang=('SL_ChuyenThang', 'sum'),
        SL_NoiBo=('SL_NoiBo', 'sum')
    ).reset_index()

    agg_status = df_valid.groupby(['Tên CHXD', 'Trạng_Thái_Lower']).size().reset_index(name='So_Luong_Trang_Thai')

    stores = sorted([s for s in df_all['Tên CHXD'].unique() if str(s).strip() != 'nan'])
    rows = []

    # TẠO TỪ ĐIỂN CỘT CHO BẢNG PHẲNG
    tuples = [('STT', ''), ('Đơn vị', '')]
    for m in metrics.keys():
        for p in products: tuples.append((m, p))
    for s in statuses:
        tuples.append(('Hóa đơn', s))

    for store in stores:
        row_data = {('STT', ''): len(rows) + 1, ('Đơn vị', ''): store}
        store_prod = agg_prod[agg_prod['Tên CHXD'] == store]
        store_status = agg_status[agg_status['Tên CHXD'] == store]
        
        for p in products:
            p_data = store_prod[store_prod['Nhóm_Hàng'] == p]
            if not p_data.empty:
                row_data[('Tổng số lượng hóa đơn', p)] = p_data['So_Luong_Dong'].iloc[0]
                row_data[('Sản lượng', p)] = p_data['Tong_San_Luong'].iloc[0]
                row_data[('Tổng số tiền chưa thuế', p)] = p_data['Tien_Chua_Thue'].iloc[0]
                row_data[('Tiền thuế', p)] = p_data['Tien_Thue'].iloc[0]
                row_data[('Tổng thanh toán', p)] = p_data['Tong_Thanh_Toan'].iloc[0]
                row_data[('Giao dịch chuyển thẳng', p)] = p_data['SL_ChuyenThang'].iloc[0]
                row_data[('Giao dịch nội bộ', p)] = p_data['SL_NoiBo'].iloc[0]
            else:
                row_data[('Tổng số lượng hóa đơn', p)] = 0
                row_data[('Sản lượng', p)] = 0
                row_data[('Tổng số tiền chưa thuế', p)] = 0
                row_data[('Tiền thuế', p)] = 0
                row_data[('Tổng thanh toán', p)] = 0
                row_data[('Giao dịch chuyển thẳng', p)] = 0
                row_data[('Giao dịch nội bộ', p)] = 0
                
        for status in statuses:
            s_data = store_status[store_status['Trạng_Thái_Lower'] == status.lower()]
            if not s_data.empty:
                row_data[('Hóa đơn', status)] = s_data['So_Luong_Trang_Thai'].iloc[0]
            else:
                row_data[('Hóa đơn', status)] = 0
            
        rows.append(row_data)

    if not rows: return None

    # CHUYỂN DỮ LIỆU THÀNH DẠNG BẢNG PHẲNG ĐỂ GHI TRỰC TIẾP
    flat_data = []
    for r in rows:
        flat_row = []
        for t in tuples:
            val = r.get(t, "")
            if val == 0: val = ""
            flat_row.append(val)
        flat_data.append(flat_row)

    result_df = pd.DataFrame(flat_data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        result_df.to_excel(writer, index=False, header=False, startrow=2, sheet_name='Tổng hợp') 
        worksheet = writer.sheets['Tổng hợp']
        
        # VẼ HEADER THỦ CÔNG & TỰ ĐỘNG GỘP Ô
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, (level1, level2) in enumerate(tuples, start=1):
            cell1 = worksheet.cell(row=1, column=col_idx, value=level1)
            cell2 = worksheet.cell(row=2, column=col_idx, value=level2)
            cell1.fill = cell2.fill = header_fill
            cell1.font = cell2.font = header_font
            cell1.alignment = cell2.alignment = center_align

        # GỘP DỌC (STT & Đơn vị)
        worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        worksheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

        # GỘP NGANG (TỰ ĐỘNG ÁP DỤNG CHO CẢ 2 NHÓM MỚI VỪA THÊM)
        start_col = 3
        for metric_group in metrics.keys():
            end_col = start_col + len(products) - 1
            worksheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            start_col = end_col + 1
            
        end_col = start_col + len(statuses) - 1
        worksheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

        # CĂN CHỈNH CỘT VÀ FORMAT SỐ NGHÌN (,000)
        worksheet.column_dimensions['A'].width = 5
        worksheet.column_dimensions['B'].width = 25
        for col_idx in range(3, len(tuples) + 1):
            worksheet.column_dimensions[get_column_letter(col_idx)].width = 15
            
        for row in range(3, worksheet.max_row + 1):
            for col_idx in range(3, len(tuples) + 1):
                cell = worksheet.cell(row=row, column=col_idx)
                if isinstance(cell.value, (int, float)) and cell.value != "":
                    cell.number_format = '#,##0'

    output.seek(0)
    return output

def generate_excel_from_bq(agg_prod: pd.DataFrame, agg_status: pd.DataFrame) -> io.BytesIO:
    """
    Hàm này thay thế nửa sau của aggregate_hd01_data cũ. 
    Nó nhận kết quả tính sẵn từ SQL BigQuery và tiến hành vẽ bảng Excel đa tầng.
    """
    if agg_prod.empty: return None

    target_products = ['Xăng RON95 Mức 3', 'Xăng E5 RON92 Mức 2', 'Dầu Điêzen 0,001S Mức 5', 'Dầu Điêzen 0,05S Mức 2', 'Xăng E10 RON95 Mức 3']
    products = target_products + ['Mặt hàng khác']
    statuses = ['Hoàn thành', 'Thay thế', 'Điều chỉnh tăng', 'Điều chỉnh giảm', 'Bị thay thế', 'Bị điều chỉnh']
    
    tuples = [('STT', ''), ('Đơn vị', '')]
    for m in ['Tổng số lượng hóa đơn', 'Sản lượng', 'Tổng số tiền chưa thuế', 'Tiền thuế', 'Tổng thanh toán', 'Giao dịch chuyển thẳng', 'Giao dịch nội bộ']:
        for p in products: tuples.append((m, p))
    for s in statuses: tuples.append(('Hóa đơn', s))

    stores = sorted(list(set(agg_prod['Ten_CHXD'].dropna().unique().tolist() + agg_status['Ten_CHXD'].dropna().unique().tolist())))
    rows = []

    for store in stores:
        row_data = {('STT', ''): len(rows) + 1, ('Đơn vị', ''): store}
        store_prod = agg_prod[agg_prod['Ten_CHXD'] == store]
        store_status = agg_status[agg_status['Ten_CHXD'] == store]
        
        for p in products:
            p_data = store_prod[store_prod['Nhom_Hang'] == p]
            if not p_data.empty:
                row_data[('Tổng số lượng hóa đơn', p)] = p_data['So_Luong_Dong'].iloc[0]
                row_data[('Sản lượng', p)] = p_data['Tong_San_Luong'].iloc[0]
                row_data[('Tổng số tiền chưa thuế', p)] = p_data['Tien_Chua_Thue'].iloc[0]
                row_data[('Tiền thuế', p)] = p_data['Tien_Thue'].iloc[0]
                row_data[('Tổng thanh toán', p)] = p_data['Tong_Thanh_Toan'].iloc[0]
                row_data[('Giao dịch chuyển thẳng', p)] = p_data['SL_ChuyenThang'].iloc[0]
                row_data[('Giao dịch nội bộ', p)] = p_data['SL_NoiBo'].iloc[0]
            else:
                row_data[('Tổng số lượng hóa đơn', p)] = 0
                row_data[('Sản lượng', p)] = 0
                row_data[('Tổng số tiền chưa thuế', p)] = 0
                row_data[('Tiền thuế', p)] = 0
                row_data[('Tổng thanh toán', p)] = 0
                row_data[('Giao dịch chuyển thẳng', p)] = 0
                row_data[('Giao dịch nội bộ', p)] = 0
                
        for status in statuses:
            s_data = store_status[store_status['Trang_Thai_Lower'] == status.lower()]
            row_data[('Hóa đơn', status)] = s_data['So_Luong_Trang_Thai'].iloc[0] if not s_data.empty else 0
            
        rows.append(row_data)

    flat_data = []
    for r in rows:
        flat_row = []
        for t in tuples:
            val = r.get(t, "")
            if val == 0: val = ""
            flat_row.append(val)
        flat_data.append(flat_row)

    result_df = pd.DataFrame(flat_data)
    output = io.BytesIO()
    
    # Mượn lại chính hàm vẽ khung Excel siêu việt của bạn
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        result_df.to_excel(writer, index=False, header=False, startrow=2, sheet_name='Tổng hợp') 
        worksheet = writer.sheets['Tổng hợp']
        
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, (level1, level2) in enumerate(tuples, start=1):
            cell1 = worksheet.cell(row=1, column=col_idx, value=level1)
            cell2 = worksheet.cell(row=2, column=col_idx, value=level2)
            cell1.fill = cell2.fill = header_fill
            cell1.font = cell2.font = header_font
            cell1.alignment = cell2.alignment = center_align

        worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        worksheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

        start_col = 3
        for _ in range(7): # 7 nhóm metrics
            end_col = start_col + len(products) - 1
            worksheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            start_col = end_col + 1
            
        end_col = start_col + len(statuses) - 1
        worksheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

        worksheet.column_dimensions['A'].width = 5
        worksheet.column_dimensions['B'].width = 25
        for col_idx in range(3, len(tuples) + 1): worksheet.column_dimensions[get_column_letter(col_idx)].width = 15
            
        for row in range(3, worksheet.max_row + 1):
            for col_idx in range(3, len(tuples) + 1):
                cell = worksheet.cell(row=row, column=col_idx)
                if isinstance(cell.value, (int, float)) and cell.value != "": cell.number_format = '#,##0'

    output.seek(0)
    return output