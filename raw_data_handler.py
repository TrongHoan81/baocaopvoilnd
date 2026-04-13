# -*- coding: utf-8 -*-
import pandas as pd
import io
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def export_to_excel_raw(df):
    """
    Hàm nguyên tử: Biến DataFrame thô từ BigQuery thành file Excel chuyên nghiệp.
    Tách biệt hoàn toàn để không ảnh hưởng đến logic tính toán tổng hợp.
    """
    if df is None or df.empty:
        return None

    output = io.BytesIO()
    # Sử dụng engine openpyxl để format bảng biểu
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Du_Lieu_Tho_HD01')
        
        workbook = writer.book
        worksheet = writer.sheets['Du_Lieu_Tho_HD01']
        
        # 1. Định dạng Header (Dòng 1)
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        
        for col_idx, cell in enumerate(worksheet[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            
            # 2. Tự động chỉnh độ rộng cột cơ bản
            column_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = 22

        # 3. Định dạng dữ liệu: Căn giữa các cột mã, ngày tháng, trạng thái
        center_cols = ['Ma_CHXD', 'Ky_Hieu', 'So_HD', 'Ngay_Hoa_Don', 'DVT', 'Trang_Thai_HD']
        for col_idx, column in enumerate(df.columns, 1):
            if column in center_cols:
                for row in range(2, worksheet.max_row + 1):
                    worksheet.cell(row=row, column=col_idx).alignment = center_align

    output.seek(0)
    return output